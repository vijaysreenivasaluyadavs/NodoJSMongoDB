import openpyxl
import requests
from pymongo import MongoClient

# Give the location of the file
path = "D:\\Vijay\\Guarantee Scheme_Scripts\\"

# Path to your Excel file
filename = 'ESCOM_Data-District.xlsx'

# To open the workbook 
# workbook object is created
wb_obj = openpyxl.load_workbook(path + filename)
ws = wb_obj.active

max_cols = ws.max_column
max_rows = ws.max_row

xls_data = []
xls_data_dict = []
for row in ws.iter_rows(values_only=True):  # Returns each row as a tuple
    xls_data.append(row)
    xls_data_dict.append({"escomId": row[0], "month": row[1], "year": row[2]})

#print ("Tuple list")
#print (xls_data)
print ("Dictionary")
del xls_data_dict[0]
print (xls_data_dict)


# API URL (Replace with the actual API endpoint)
BESCOM_api_url = "https://bescom.co.in:8081/bescom/api/v1/cm/getCmDashboardData"  # Example API

headers = {
    "Content-Type": "application/json"
    #"Authorization": "Bearer YOUR_ACCESS_TOKEN"  # Replace if required
}


# Step 1: Connect to MongoDB (Local or Remote)
connection_string = "mongodb+srv://vijayyadavs:4idF7xjM7ni8h1Aq@sample.sc7oj2n.mongodb.net/"
client = MongoClient(connection_string)  # Use your MongoDB connection string

# Step 2: Select Database
db = client["MySampleProject"]  # Change to your database name

# Step 3: Select Collection (Table)
collection = db["Vijay_Collection"]  # Change to your collection name

#for row in ws.iter_rows(values_only=True):  # Returns each row as a tuple
# Request Payload (Body)
for item in xls_data_dict:
    print ("item")
    print (item)
    data = item

    # Sending the POST request
    response = requests.post(BESCOM_api_url, json=data, headers=headers)

    # Check if the request was successful
    if response.status_code == 200:  # 200 Created (successful POST)
        print ("Received Successfully from BESCOM")
        print (response.status_code),
        print("Successfully retrieved from BESCOM web API Response JSON:", response.json())  # Print response as JSON
        #{"escomId": escom_Value, "month": month_value, "year": year_value}
        
        # Step 4: Insert Single Document
        for item in response.json():
            document = item
            insert_result = collection.insert_one(document)
            print ("Something insert")
            print (insert_result)

    else:
        print(f"Error: {response.status_code}, {response.text}")  # Print error message

# Step 6: Verify Data (Optional)
for doc in collection.find():
    print(doc)

# Step 7: Close the Connection (Optional)
client.close()
